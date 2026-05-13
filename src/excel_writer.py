from __future__ import annotations

"""Запись полимерных уведомлений в Excel-файлы.

Каждый Excel-файл сопровождается JSON-кэшем ID, которые были когда-либо
записаны.  Если пользователь удалил строку из Excel вручную, ID остаётся
в кэше и строка не будет добавлена повторно.
"""

import json
import logging
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.dashboard_config import DATA_DIR

logger = logging.getLogger(__name__)

# Пути к файлам
SP_EXCEL_PATH = os.path.normpath(os.path.join(DATA_DIR, "polymer_sp_registry.xlsx"))
SP_IDS_CACHE = os.path.normpath(os.path.join(DATA_DIR, "polymer_sp_written_ids.json"))

GOST_EXCEL_PATH = os.path.normpath(os.path.join(DATA_DIR, "polymer_gost_registry.xlsx"))
GOST_IDS_CACHE = os.path.normpath(os.path.join(DATA_DIR, "polymer_gost_written_ids.json"))

# Стили
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL_SP = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_FILL_GOST = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ------------------------------------------------------------------
# Кэш ID (JSON)
# ------------------------------------------------------------------
def _load_written_ids(cache_path: str) -> set[str]:
    """Загружает множество ID, которые были записаны в Excel."""
    if not os.path.exists(cache_path):
        return set()
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return set(data.get("written_ids", []))
    except (json.JSONDecodeError, OSError):
        return set()


def _save_written_ids(cache_path: str, ids: set[str]) -> None:
    """Сохраняет множество записанных ID."""
    os.makedirs(os.path.dirname(cache_path), exist_ok=True)
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump({"written_ids": sorted(ids)}, f, ensure_ascii=False, indent=2)


# ------------------------------------------------------------------
# Создание / открытие Excel
# ------------------------------------------------------------------
def _get_or_create_workbook(path: str, headers: list[str], fill) -> Workbook:
    """Открывает существующий Excel или создаёт новый с заголовками."""
    if os.path.exists(path):
        return load_workbook(path)

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр"

    # Записываем заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = _HEADER_ALIGN

    # Скрываем столбец ID (первый)
    ws.column_dimensions["A"].hidden = True

    # Замораживаем заголовок
    ws.freeze_panes = "A2"

    return wb


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Подгоняет ширину столбцов под содержимое (кроме скрытого ID)."""
    for col_idx in range(2, ws.max_column + 1):  # пропускаем столбец A (ID)
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


# ------------------------------------------------------------------
# СП (Своды правил)
# ------------------------------------------------------------------
_SP_HEADERS = [
    "ID",
    "Дата добавления",
    "Тип уведомления",
    "Наименование проекта",
    "Тип документа",
    "Разработчик",
    "Дата размещения",
    "Ключевые слова",
    "Ссылка",
]


def update_sp_excel(notifications: list[dict]) -> None:
    """Добавляет полимерные СП-уведомления в Excel-реестр."""
    if not notifications:
        return

    written_ids = _load_written_ids(SP_IDS_CACHE)
    new_items = [n for n in notifications if n.get("id") and n["id"] not in written_ids]

    if not new_items:
        logger.info("Excel СП: нет новых полимерных записей для добавления.")
        return

    wb = _get_or_create_workbook(SP_EXCEL_PATH, _SP_HEADERS, _HEADER_FILL_SP)
    ws = wb.active
    today = datetime.now().strftime("%d.%m.%Y")

    for n in new_items:
        ws.append([
            n.get("id", ""),
            today,
            n.get("notification_type", ""),
            n.get("project_name", n.get("title", "")),
            n.get("doc_type", "Свод правил"),
            n.get("developer", ""),
            n.get("placement_date", n.get("date", "")),
            ", ".join(n.get("matched_keywords", [])),
            n.get("url", ""),
        ])
        written_ids.add(n["id"])

    _auto_width(ws)
    wb.save(SP_EXCEL_PATH)
    _save_written_ids(SP_IDS_CACHE, written_ids)
    logger.info(f"Excel СП: добавлено {len(new_items)} записей → {SP_EXCEL_PATH}")


# ------------------------------------------------------------------
# ГОСТ
# ------------------------------------------------------------------
_GOST_HEADERS = [
    "ID",
    "Дата добавления",
    "Тип документа",
    "Наименование проекта стандарта",
    "Шифр ПНС",
    "Технический комитет",
    "Разработчик",
    "Дата начала обсуждения",
    "Дата завершения обсуждения",
    "Статус",
    "Ключевые слова",
    "Ссылка",
    "Комментарий направлен",
]


def _migrate_gost_add_comment_column(path: str) -> None:
    """Добавляет столбец 'Комментарий направлен' если его нет в существующем файле."""
    if not os.path.exists(path):
        return
    wb = load_workbook(path)
    ws = wb.active
    # Проверяем заголовок последнего столбца
    last_col = ws.max_column
    if last_col < 13 or ws.cell(row=1, column=13).value != "Комментарий направлен":
        col_idx = 13  # Столбец M
        cell = ws.cell(row=1, column=col_idx, value="Комментарий направлен")
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL_GOST
        cell.alignment = _HEADER_ALIGN
        wb.save(path)
        logger.info(f"Миграция Excel ГОСТ: добавлен столбец «Комментарий направлен» → {path}")


def read_polymer_gost_stats() -> dict:
    """Читает Excel-реестр полимерных ГОСТов и возвращает статистику.

    Returns:
        dict с ключами:
        - total: общее количество полимерных ГОСТов
        - commented: количество с направленными комментариями
    """
    result = {"total": 0, "commented": 0}

    if not os.path.exists(GOST_EXCEL_PATH):
        return result

    wb = load_workbook(GOST_EXCEL_PATH, read_only=True)
    ws = wb.active

    # Ищем столбец "Комментарий направлен"
    comment_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == "Комментарий направлен":
            comment_col = col_idx
            break

    for row_idx in range(2, ws.max_row + 1):
        # Проверяем что строка не пустая (есть ID в столбце A)
        if not ws.cell(row=row_idx, column=1).value:
            continue
        result["total"] += 1
        if comment_col:
            val = ws.cell(row=row_idx, column=comment_col).value
            if val and str(val).strip():
                result["commented"] += 1

    wb.close()
    return result


def update_gost_excel(notifications: list[dict]) -> None:
    """Добавляет полимерные ГОСТ-уведомления в Excel-реестр."""
    if not notifications:
        return

    written_ids = _load_written_ids(GOST_IDS_CACHE)
    new_items = [n for n in notifications if n.get("id") and n["id"] not in written_ids]

    if not new_items:
        logger.info("Excel ГОСТ: нет новых полимерных записей для добавления.")
        return

    # Миграция: добавляем столбец «Комментарий направлен» если его нет
    _migrate_gost_add_comment_column(GOST_EXCEL_PATH)

    wb = _get_or_create_workbook(GOST_EXCEL_PATH, _GOST_HEADERS, _HEADER_FILL_GOST)
    ws = wb.active
    today = datetime.now().strftime("%d.%m.%Y")

    for n in new_items:
        ws.append([
            n.get("id", ""),
            today,
            n.get("doc_type", ""),
            n.get("project_name", ""),
            n.get("prns_code", ""),
            n.get("technical_committee", ""),
            n.get("developer", ""),
            n.get("start_date", ""),
            n.get("end_date", ""),
            n.get("status", ""),
            ", ".join(n.get("matched_keywords", [])),
            n.get("url", ""),
            "",  # Комментарий направлен (заполняется пользователем вручную)
        ])
        written_ids.add(n["id"])

    _auto_width(ws)
    wb.save(GOST_EXCEL_PATH)
    _save_written_ids(GOST_IDS_CACHE, written_ids)
    logger.info(f"Excel ГОСТ: добавлено {len(new_items)} записей → {GOST_EXCEL_PATH}")
